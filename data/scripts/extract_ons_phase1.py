#!/usr/bin/env python3
"""Extract ONS regional finance and council-region lookup for Phase 1."""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
RAW = ROOT / "data" / "raw"
OUT = ROOT / "data" / "processed"


GEOGRAPHY_CODE_BY_SHEET = {
    "United Kingdom": "K02000001",
    "North East": "E12000001",
    "North West": "E12000002",
    "Yorkshire and The Humber": "E12000003",
    "East Midlands": "E12000004",
    "West Midlands": "E12000005",
    "East of England": "E12000006",
    "London": "E12000007",
    "South East": "E12000008",
    "South West": "E12000009",
    "England": "E92000001",
    "Wales": "W92000004",
    "Scotland": "S92000003",
    "Northern Ireland": "N92000002",
}


def find_latest_year_col(ws, header_row: int = 3) -> tuple[str, int]:
    latest_col = None
    latest_year = None
    for col in range(2, ws.max_column + 1):
        value = ws.cell(header_row, col).value
        if isinstance(value, str) and "to" in value:
            latest_col = col
            latest_year = value
    if latest_col is None or latest_year is None:
        raise ValueError(f"Could not find latest year header in sheet {ws.title}")
    return latest_year, latest_col


def find_row_by_prefix(ws, prefix: str) -> int:
    wanted = prefix.strip().lower()
    for row in range(1, ws.max_row + 1):
        value = ws.cell(row, 1).value
        if isinstance(value, str) and value.strip().lower().startswith(wanted):
            return row
    raise ValueError(f"Could not find row with prefix '{prefix}' in sheet {ws.title}")


def extract_regional_revenue() -> Path:
    wb = load_workbook(RAW / "ons_regional_revenue_fye2023.xlsx", read_only=True, data_only=True)
    rows: list[dict[str, object]] = []
    for sheet_name, geography_code in GEOGRAPHY_CODE_BY_SHEET.items():
        ws = wb[sheet_name]
        year_label, latest_col = find_latest_year_col(ws, header_row=3)
        data_row = find_row_by_prefix(
            ws, "Total current receipts (excl. North Sea Oil & Gas revenues)"
        )
        amount = ws.cell(data_row, latest_col).value
        rows.append(
            {
                "year": year_label,
                "geography_code": geography_code,
                "geography_name": sheet_name,
                "metric": "total_current_receipts_excl_north_sea_oil_gas",
                "amount_m_gbp": float(amount),
                "source_table": "ONS_CRPSF_Revenue_FYE2023",
            }
        )

    out_path = OUT / "ons_regional_revenue_fye2023.csv"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "year",
                "geography_code",
                "geography_name",
                "metric",
                "amount_m_gbp",
                "source_table",
            ],
        )
        writer.writeheader()
        writer.writerows(rows)
    return out_path


def extract_regional_expenditure() -> Path:
    wb = load_workbook(
        RAW / "ons_regional_expenditure_fye2023.xlsx", read_only=True, data_only=True
    )
    rows: list[dict[str, object]] = []
    for sheet_name, geography_code in GEOGRAPHY_CODE_BY_SHEET.items():
        ws = wb[sheet_name]
        year_label, latest_col = find_latest_year_col(ws, header_row=3)
        data_row = find_row_by_prefix(ws, "Total managed expenditure")
        amount = ws.cell(data_row, latest_col).value
        rows.append(
            {
                "year": year_label,
                "geography_code": geography_code,
                "geography_name": sheet_name,
                "metric": "total_managed_expenditure",
                "amount_m_gbp": float(amount),
                "source_table": "ONS_CRPSF_Expenditure_FYE2023",
            }
        )

    out_path = OUT / "ons_regional_expenditure_fye2023.csv"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "year",
                "geography_code",
                "geography_name",
                "metric",
                "amount_m_gbp",
                "source_table",
            ],
        )
        writer.writeheader()
        writer.writerows(rows)
    return out_path


def extract_council_to_region_lookup() -> Path:
    in_path = RAW / "ons_lad_to_region_2024.csv"
    out_path = OUT / "council_to_region_2024.csv"
    rows = []
    with in_path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append(
                {
                    "lad_code": row["LAD24CD"].strip(),
                    "lad_name": row["LAD24NM"].strip(),
                    "region_code": row["RGN24CD"].strip(),
                    "region_name": row["RGN24NM"].strip(),
                }
            )
    rows.sort(key=lambda r: r["lad_code"])
    with out_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f, fieldnames=["lad_code", "lad_name", "region_code", "region_name"]
        )
        writer.writeheader()
        writer.writerows(rows)
    return out_path


def main() -> None:
    revenue = extract_regional_revenue()
    expenditure = extract_regional_expenditure()
    lookup = extract_council_to_region_lookup()
    print(f"Wrote {revenue}")
    print(f"Wrote {expenditure}")
    print(f"Wrote {lookup}")


if __name__ == "__main__":
    main()
