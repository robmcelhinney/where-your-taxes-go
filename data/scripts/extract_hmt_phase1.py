#!/usr/bin/env python3
"""Extract Phase 1 HMT spending tables into normalized CSV snapshots.

Run with:
  uv run --with openpyxl python data/scripts/extract_hmt_phase1.py
"""

from __future__ import annotations

import csv
import re
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
RAW = ROOT / "data" / "raw"
OUT = ROOT / "data" / "processed"


def to_number(value):
    if value in (None, "", "-"):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").strip()
    if text == "-" or text == "":
        return None
    try:
        return float(text)
    except ValueError:
        return None


def classify_function_label(label: str) -> str:
    if label.lower().startswith("of which"):
        return "of_which"
    if label.lower().startswith("total "):
        return "function_total"
    if re.match(r"^\d+\.\d+\s", label):
        return "sub_function"
    if re.match(r"^\d+\.\s", label):
        return "function_heading"
    return "other"


def extract_functional_categories() -> Path:
    wb = load_workbook(RAW / "pesa_2025_ch5_tables.xlsx", data_only=True, read_only=True)
    ws = wb["5_2"]
    rows = []
    for r in range(1, ws.max_row + 1):
        label = ws.cell(r, 2).value
        amount = to_number(ws.cell(r, 7).value)  # 2024-25 outturn
        if not label or amount is None:
            continue
        rows.append(
            {
                "year": "2024-25",
                "source_table": "PESA_2025_5_2",
                "function_label": str(label).strip(),
                "row_type": classify_function_label(str(label).strip()),
                "amount_m_gbp": amount,
            }
        )

    out_path = OUT / "functional_spending_2024_25.csv"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f, fieldnames=["year", "source_table", "function_label", "row_type", "amount_m_gbp"]
        )
        writer.writeheader()
        writer.writerows(rows)
    return out_path


def classify_department_label(label: str) -> str:
    lower = label.lower()
    if lower.startswith("total "):
        return "total"
    if lower.startswith("("):
        return "note"
    return "line_item"


def extract_departmental_totals() -> Path:
    wb = load_workbook(
        RAW / "public_spending_departmental_budgets_july_2025.xlsx",
        data_only=True,
        read_only=True,
    )
    ws = wb["Table_1_12"]
    rows = []
    for r in range(1, ws.max_row + 1):
        label = ws.cell(r, 1).value
        amount = to_number(ws.cell(r, 6).value)  # 2024-25 outturn
        if not label or amount is None:
            continue
        label = str(label).strip()
        rows.append(
            {
                "year": "2024-25",
                "source_table": "PSS_July_2025_Table_1_12",
                "department_label": label,
                "row_type": classify_department_label(label),
                "amount_m_gbp": amount,
            }
        )

    out_path = OUT / "departmental_spending_2024_25.csv"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f, fieldnames=["year", "source_table", "department_label", "row_type", "amount_m_gbp"]
        )
        writer.writeheader()
        writer.writerows(rows)
    return out_path


def main() -> None:
    functional = extract_functional_categories()
    departmental = extract_departmental_totals()
    print(f"Wrote {functional}")
    print(f"Wrote {departmental}")


if __name__ == "__main__":
    main()
